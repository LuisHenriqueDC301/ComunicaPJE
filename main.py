from datetime import datetime, timedelta
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from tkinter import Tk, filedialog
import re

Tk().withdraw()

wb = Workbook()
ws = wb.active

cabecalho_font = Font(bold=True, color="000000")
cabecalho_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

cabecalhos = ["ADVOGADO/OAB", "N.º PROCESSO", "SIGLA TRIBUNAL", "DATA DE DISPONIBILIZAÇÃO"]
ws.append(cabecalhos)

for col_num, titulo in enumerate(cabecalhos, start=1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = cabecalho_font
    cell.fill = cabecalho_fill
    ws.column_dimensions[cell.column_letter].width = len(titulo) + 2

lista_oab = [169990, 78115, 164700, 74792, 208355, 177742, 56154, 176082, 107733, 157985, 141697]
data_inicio = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
nome_parte = "EMATER-MG"

url = (
    f"https://comunicaapi.pje.jus.br/api/v1/comunicacao?"
    f"nomeParte={nome_parte}&dataDisponibilizacaoInicio={data_inicio}&"
    f"dataDisponibilizacaoFim=2025-07-18&pagina=1&itensPorPagina=100"
)

try:
    response = requests.get(url)
    response.raise_for_status()
    json_response = response.json()

    for processo in json_response.get("items", []):
        numero_processo = processo["numero_processo"]
        sigla_tribunal = processo.get("siglaTribunal", "N/A")
        data_disponibilizacao = processo["datadisponibilizacao"]
        destinatario_advogados = processo.get("destinatarioadvogados", [])

        lista_advogados = set()
        for destinatario in destinatario_advogados:
            advogado_info = destinatario["advogado"]
            numero_oab = advogado_info["numero_oab"]
            nome_advogado = advogado_info["nome"]
            uf_oab = advogado_info["uf_oab"]

            numero_oab_limpo = re.sub(r'\D', '', numero_oab)

            if numero_oab_limpo and int(numero_oab_limpo) in lista_oab:
                advogado_completo = f"{nome_advogado} - {numero_oab}/{uf_oab}"
                lista_advogados.add(advogado_completo)

        if lista_advogados:
            for advogado in lista_advogados:
                nova_linha = [advogado, numero_processo, sigla_tribunal, data_disponibilizacao]
                ws.append(nova_linha)

            # Linha em branco para separar processos
            ws.append([''] * len(cabecalhos))

            print(f"Processo: {numero_processo} | Tribunal: {sigla_tribunal} | Advogados: {', '.join(lista_advogados)}")

        else:
            nova_linha = ["Nenhum advogado na lista", numero_processo, sigla_tribunal, data_disponibilizacao]
            ws.append(nova_linha)
            ws.append([''] * len(cabecalhos))
            print(f"Processo: {numero_processo} | Tribunal: {sigla_tribunal} | Nenhum advogado encontrado.")

    # Ajusta largura das colunas com base no conteúdo
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    nome_sugerido = f"planilha_emater_{data_inicio}.xlsx"
    caminho_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        initialfile=nome_sugerido,
        title="Salvar planilha como"
    )

    if caminho_arquivo:
        wb.save(caminho_arquivo)
        print(f"Planilha salva com sucesso em: {caminho_arquivo}")
    else:
        print("Operação de salvamento cancelada pelo usuário.")

except requests.exceptions.HTTPError as err:
    print(f"Erro HTTP: {err}")

except Exception as e:
    print(f"Erro inesperado: {e}")
